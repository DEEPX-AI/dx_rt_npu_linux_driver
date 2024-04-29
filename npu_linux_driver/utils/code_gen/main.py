## Common Python module ##
from fileinput import filename
import inspect
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from copy import copy
import sympy
from mako.template import Template
import argparse
import re
from pathlib import Path

## User Python module ##
from config import *

### Target System ###
ASIC   = 0
FPGA   = 1
TARGET = ["ASIC", "FPGA"]

## Global Variable ##
BASE_ADDR       = 0
SPLIT_FILE_N    = "split.xlsx"

parser = argparse.ArgumentParser(description='Generate PCIe register and cmm scripts')
parser.add_argument('--target', type=str, default=TARGET[ASIC],
                    help=f"[ASIC, FPGA] \
                        This parameters will be effected to diag_sts_bus base address \n \
                        (ASIC:0x404, FPGA:{DIAG_STS_BUS_OFFSET})")
parser.add_argument('--dev', type=int, default=DX_M1A,
                    help=f'target device : {DEV_T}')

pcie_conf_value = dict()
target_env      = ASIC
target_sys      = ""
g_dev           = ""
generated_f     = ""

# return value : start,end or start,None
def get_bit_position(str):
    pattern = r"\[(\d+)(?::(\d+))?\]"
    parsed = re.search(pattern, str)
    return [parsed.group(1), parsed.group(2)]
def set_pcie_conf_value():
    global pcie_conf_value
    assert g_dev!="", "Please check global device name!!"
    if target_env == FPGA:
        pcie_conf_value = PCIE_CONFIG_V_FPGA
    else:
        pcie_conf_value = PCIE_CONFIG_V[g_dev]
def replace_bit(str):
    global pcie_conf_value
    for conf, v in pcie_conf_value.items():
        if conf in str:
            str = str.replace(conf, v)
    return str
def calculate_bit(str):
    calc = sympy.sympify(str)
    # print(f"Calculated : {str} -> {calc}")
    return calc
# remove ':' charecters
def removeDuplicateLetters(s):
    f_ch   = ":"
    result = s
    s_cnt  = s.count(f_ch)
    if s_cnt > 1:
        result = result.replace(f_ch, '', s_cnt-1)
    return result
def get_hif_code_lists(lists):
    code_lists = []
    for l in lists:
        if l[1]:
            code_lists.append(l[0])
    return code_lists

D_BIT_POS    = 0
D_SIZE       = 1
D_DESC       = 3
D_NAME       = 4
class Reg:
    def __init__(self):
        self.name           = list()
        self.offs           = ""
        self.size           = list()
        self.bits           = list()
        self.width          = list()
        self.mask           = list()
        self.default        = list()
        self.type           = list()
        self.desc           = list()
        self.has_desc_info  = list()

    def show(self):
        print(f"Name    : {self.name}")
        print(f"offs    : {hex(self.offs)}")
        print(f"size    : {self.size}")
        print(f"Bit     : {self.bits}")
        print(f"Default : {self.default}")
        print(f"type    : {self.type}")
        print(f"has_desc_info : {self.has_desc_info}")
        # In case that Bit information is existed in description filed. [Position, Size, Masking, Description, Name]
        if np.array(self.desc[0], dtype=object).ndim > 0:
            for idx, desc in enumerate(self.desc[0]):
                # print(f"desc[#{idx}] {desc}}")
                print(f"desc[#{idx}] Bit Pos:{desc[D_BIT_POS]}, Size:{desc[D_SIZE]}, Desc:{desc[D_DESC]}, DescName:{desc[D_NAME]}")
        else:
            print(f"desc    : {self.desc}")
        print(f"desc dimension   : {np.array(self.desc, dtype=object).shape}")

class RegList:
    def __init__(self):
        self.r_lists = dict()

    def exist(self, r_name):
        ret = False
        if r_name in self.r_lists:
            ret = True
        return ret

    def exist_name(self, module, r_name):
        ret = False
        exist = self.r_lists.get(module)
        if exist:
            if r_name in self.r_lists[module].name:
                ret = True
        return ret

    def desc_checker(self):
        ret = False
        for l in list:
            if np.array(l).shape:
                ret = True
        return ret

    ## Remove white space of descriotion in register
    def remove_white_space_of_desc(self):
        for module, regs in self.r_lists.items():
            if regs.has_desc_info:
                for idx, desc in enumerate(regs.desc):
                    regs.desc[idx] = str(desc).replace("\n", " \\ ").replace("\r\n", " \\ ").replace("\"", "'").strip()

    def remove_white_space_of_desc_normal(self):
        for module, regs in self.r_lists.items():
            # print(f'Name : {regs.name}')
            for idx in range(0, len(regs.name)):
                if not regs.has_desc_info[idx]:
                    regs.desc[idx] = str(regs.desc[idx]).replace("\n", " \\ ").replace("\r\n", " \\ ").replace("\"", "'").strip()
                else:
                    for idx, desc in enumerate(regs.desc[idx]):
                        desc[D_DESC] = desc[D_DESC].replace('"','')
                        # print(f'#{idx}, {desc[D_DESC]}')

    # return : [bit position, description]
    def decode_desc(self, data, reg_n):
        global pcie_conf_value
        ret = []
        skip = 0
        for d in str(data).split('\n'):
            ## if RX_TLP is '1' then skip
            if pcie_conf_value['RX_TLP'] == "1" and DIAG_STS_BUS in reg_n:
                skip_s, skip_e = "RADM Diagnostic Bus 1", "XADM Diagnostic Bus"
                if skip_s == d.strip():
                    skip = 1
                if skip_e == d.strip():
                    skip = 0
                if skip == 1:
                    continue
            ## Parsing datas
            if '■' in d:
                line = d.replace('■', '')
                #print(f"line : {line}")
                if DIAG_STS_BUS in reg_n:
                    c_pos_s = line.find('[')
                    c_pos_e = line.find(']')+1
                    if (c_pos_s == -1) or (line[c_pos_s:c_pos_e].find(":") == -1):
                        c_pos_e = line.find('=')
                        ret.append(["[0]", line[c_pos_e:]])
                    else:
                        ret.append([line[c_pos_s:c_pos_e], line[c_pos_e:]])
                    ## register name of desc
                    r_name_idx = line.find('[')
                    r_name_sep = line.find('=')
                    if ('[' in line) and (r_name_idx < r_name_sep):
                        ret[len(ret)-1].append(line.split('[')[0].strip())
                    else:
                        ret[len(ret)-1].append(line.split('=')[0].strip())
                else:
                    # check whether reg name is null or not
                    r_name_idx = line.strip().find('[')
                    c_pos = line.find(']')+1
                    if r_name_idx == 0:
                        ## [] : xxxxx : yyyy -> In case that "xxxxx : yyyy" is description
                        first_char_idx = line.find(':', c_pos)
                        second_char_idx = line.find(':', first_char_idx+1)
                        reg_name_valid_chk = line[first_char_idx+1:second_char_idx].strip().find(' ')
                        print(f"second_char_idx: {second_char_idx}, reg_name_valid_chk:{reg_name_valid_chk}, {line[first_char_idx+1:second_char_idx]}")
                        if second_char_idx != -1 and reg_name_valid_chk == -1:
                            ## Bit Position, Description
                            ret.append([line[:c_pos], line[first_char_idx:]])
                        else:
                            ret.append([line[:c_pos], line[c_pos:], "bit"+get_bit_position(line)[0]])
                    else:
                        ret.append([line[:c_pos], line[c_pos:]])
        if ret:
            data = ret
            # print("*Decoded return data:", ret)
        return data

    def diag_decode_desc(self):
        print("Start Diag Description Decodeing!!")
        diag_r = self.r_lists[DIAG_STS_BUS]
        bit_s, bit_e = 0, -1
        for idx, l in enumerate(diag_r.desc[0]):
            bit = replace_bit(l[0])
            # print(bit)
            # Multi-bit
            sep = bit.find(':')
            if sep != -1:
                e = calculate_bit( bit[1:sep] )
                if (int(e)-0) > 0:
                    bit_s = (bit_e + 1)
                    bit_e = int(e) + (bit_e + 1)
                    diag_r.desc[0][idx][0] = "[{}:{}]".format(bit_e, bit_s)
                else:
                    bit_s += 1
                    bit_e += 1
                    diag_r.desc[0][idx][0] = "[{}]".format(bit_e)
                # print(l)
            else:
                bit_s += 1
                bit_e += 1
                diag_r.desc[0][idx][0] = "[{}]".format(bit_e)

    def append(self, data):
        reg_t = Reg()
        r_n = data[1].strip() # Register Name
        ## in case of register name is empty
        if r_n == '':
            return
        ## check wether register name with '[' or ']'
        name = data[5]
        if '[' in name:
            name = name[:name.find('[')]

        if self.exist(r_n):
            self.r_lists[r_n].bits.append(data[4])
            self.r_lists[r_n].name.append(name.strip())
            try:
                self.r_lists[r_n].default.append(int(data[6], 16))
            except:
                self.r_lists[r_n].default.append(data[6])
            self.r_lists[r_n].type.append(data[7])
            self.r_lists[r_n].desc.append(self.decode_desc(data[9], r_n))
        else:
            if r_n == DIAG_STS_BUS and target_sys == TARGET[FPGA]:
                reg_t.offs      = int(DIAG_STS_BUS_OFFSET[DEV_T[target_dev]], 16)   # FPGA Bit File
            else:
                try:
                    if "'h" in data[2]:
                        data[2] = data[2].replace("'h", "")
                    reg_t.offs      = int(data[2], 16)
                except:
                    print(f"[ERR] Conversion to offset integer failed :{data[2]}")
                    print(f"Data : {data}")
                    exit()
            reg_t.bits.append(data[4])
            reg_t.name.append(name.strip())
            try:
                reg_t.default.append(int(data[6], 16))
            except:
                reg_t.default.append(data[6])
            reg_t.type.append(data[7].strip())
            reg_t.desc.append(self.decode_desc(data[9], r_n))
            # Insert Reg() to lists
            self.r_lists[r_n] = reg_t
            print(f"List is appended with initial Key : {r_n}")
        ## In case of DIAG_STS_BUS ##
        if DIAG_STS_BUS in r_n:
            self.diag_decode_desc()

    def show(self):
        for k, v in self.r_lists.items():
            print(f"{k} : {v.show()}")

class ScriptMaker:
    def __init__(self):
        self.datas = ""
        self.RegList = RegList()

    def reg_list_clear(self):
        self.RegList = None
        self.RegList = RegList()

    def split_merged_cells(self, input, output):
        wb = load_workbook(filename = input, data_only = True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            mcr_coord_list = [mcr.coord for mcr in sheet.merged_cells.ranges]

            for mcr in mcr_coord_list:
                min_col, min_row, max_col, max_row = range_boundaries(mcr)
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                top_left_cell_alignment = sheet.cell(row=min_row, column=min_col).alignment
                top_left_cell_format = sheet.cell(row=min_row, column=min_col).number_format
                sheet.unmerge_cells(mcr)
                for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                    for cell in row:
                        cell.value = top_left_cell_value
                        cell.alignment = copy(top_left_cell_alignment)
                        cell.number_format = copy(top_left_cell_format)
        wb.save(output)
        print("EXCEL SPLIT DONE!!")

    def create_split_datas(self, file_n):
        self.split_merged_cells(file_n, SPLIT_FILE_N)

    def get_datas(self, sheet_name='pcie_ctrl_4lane_hif_register'):
        self.datas = pd.read_excel(
                    SPLIT_FILE_N,
                    sheet_name=sheet_name, 
                    header=3,
                    keep_default_na=False,
                    dtype={
                        'Dummy'         : str , #Index-0
                        'Name'          : str , #Index-1
                        'Offset'        : str , #Index-2
                        'Reset Value'   : str , #Index-3
                        'Bit'           : str , #Index-4
                        'Bitfield'      : str , #Index-5
                        'Reset Value'   : str , #Index-6
                        'Accessable'    : str , #Index-7
                        'Attributes'    : str , #Index-8
                        'Description'   : str , #Index-9
                        }
                ).values.tolist()

    def set_reg_offset_recalculation(self, mul):
        for data in self.datas:
            data[2] = hex(int(data[2], 16) * mul)

    def convert_bit_to_pos_size_mask(self, bit, s_pos=0):
        pos_s, size, mask = 0, 0, 0
        bit = bit.strip()
        bit_s = bit.find(':')
        if bit_s != -1: # Multi-bit
            pos_s = int(bit[bit_s+1 : len(bit)-1]) - s_pos
            pos_e = int(bit[1 : bit_s]) - s_pos
            size  = pos_e - pos_s + 1
        else:           # Single-bit
            pos_s = int(bit[1:len(bit)-1]) - s_pos
            size  = 1
        for idx in range(0, size):
            mask |= (0x1<<(int(pos_s)+idx))
        return str(pos_s), str(size), str(hex(mask))

    def parsing_variable_name(self, description, pos):
        name, name_cvrt = description.replace(' ', ''), ""
        if description.find("Reserved") >= 0:
            name_cvrt = "Reserved_"+pos
        else:
            pos_s = 0
            ## Array Bit '[:]' or '[]'
            if name.find('[') != -1:
                ## '[:]'
                s = name.find('[')
                e = name.find(']')
                sep = name[s:e].find(':')
                if sep != -1:
                    name = name[:s] + "_" + name[s:e+1].replace(':','_').strip()
            pos_e = name.find(":")
            name = name.replace('[','').replace(']','').strip()
            ## Skip if ':' is 0 
            if pos_e == 0:
                pos_s = 1
                pos_e = name.find(":", pos_s)
            if name.find("*") != -1: 
                pos_e = name.find("*")
                name_cvrt = name[pos_s:pos_e].strip() + "_" + pos
            else:
                name_cvrt = name[pos_s:pos_e].strip()
            ## Remove'|'
            name_cvrt = name_cvrt.replace('|','').strip()
            # print(f"{name} -> {name_cvrt}")
        return name_cvrt

    def check_start_position(self, descs):
        ## Check if the starting position of register is 0
        desc1, desc2 = descs[0][0].strip(), descs[-1][0].strip()
        try:
            if desc1.find(":") != -1:
                start_p1 = int(desc1[desc1.find(":")+1 : len(desc1)-1])
            else:
                start_p1 = int(desc1[1 : len(desc1)-1])
            if desc2.find(":") != -1:
                start_p2 = int(desc2[desc2.find(":")+1 : len(desc2)-1])
            else:
                start_p2 = int(desc2[1 : len(desc2)-1])
        except:
            print(f"[ERR][{inspect.currentframe().f_code.co_name}] Fail: Converting to integer")
            print(f"{descs}")
            exit()
        if start_p1 == 0 or start_p2 == 0:
            start = 0
        else:
            start = start_p2 - start_p1 + 1
            if start_p1 > start_p2:
                start = start_p1 - start_p2 + 1
        return start

    ## Convert data to make script easily
    def convert_datas(self, list_reverse):
        for module, regs in self.RegList.r_lists.items():
            # print(f"Moduele : {module}")
            ## Mask of Reg
            cvrt_bit_list = []
            for bits in regs.bits:
                regs.bits, size, mask = self.convert_bit_to_pos_size_mask(bits)
                cvrt_bit_list.append(regs.bits)
                regs.size.append(size)
                regs.mask.append(mask)
            regs.bits = list.copy(cvrt_bit_list)
            ## replace register name with "RESERVED"
            for idx, name in enumerate(regs.name):
                if "RESERVED" in name.upper():
                    regs.name[idx] = "_".join([ name, regs.bits[idx] ])
            ## Mask of Description [Position, Size, Masking, Description, Variable name, struct num, struct pos, struct size]
            for idx, r_descs in enumerate(regs.desc):
                if np.array(r_descs, dtype=object).ndim > 0:
                    print(f"[{module}][{regs.name[idx]}] Convert Description is founded!!")
                    if int(regs.size[idx]) == 1: continue
                    strt_pos_desc = self.check_start_position(r_descs)
                    regs.has_desc_info.append(True)
                    for r_desc in r_descs:
                        #print(f"{r_desc[0]},  {strt_pos_desc}")
                        r_desc[0], size, mask = self.convert_bit_to_pos_size_mask(removeDuplicateLetters(r_desc[0]), strt_pos_desc)
                        r_desc.insert(1, size)
                        r_desc.insert(2, mask)
                        ## Append the variable name from description[3] to the end of the array
                        r_desc.append(self.parsing_variable_name(r_desc[3], r_desc[0]))
                else:
                    regs.has_desc_info.append(False)
                    continue
        ## Make description/register lists to ascending order based on bit position.
        for module, regs in self.RegList.r_lists.items():
            for idx, r_descs in enumerate(regs.desc):
                if np.array(r_descs).ndim > 0:
                    # Assume Description is consist of 2-demention[X, 4]
                    last_idx = np.array(r_descs).shape[0]-1
                    if (r_descs[last_idx][0] < r_descs[0][0]):
                        r_descs.reverse()
            if False in regs.has_desc_info and list_reverse:
                regs.name.reverse()
                regs.size.reverse()
                regs.bits.reverse()
                regs.mask.reverse()
                regs.default.reverse()
                regs.type.reverse()
                regs.desc.reverse()
                regs.has_desc_info.reverse()

    # Split register list if size exceeds 32 bits
    def split_register_list(self):
        for module, regs in self.RegList.r_lists.items():
            for idx in range(0, len(regs.name)):
                if regs.has_desc_info[idx]:
                    split = 0
                    for r_idx, r_desc in enumerate(regs.desc[idx]):
                        #print(f"split :: {r_desc}")
                        offs = int(r_desc[0])
                        size = int(r_desc[1])
                        # check size
                        if size > 32:
                            # calculate size and offset
                            regs_org = r_desc.copy()
                            r_desc[1] = 32-offs%32 if split==0 else 32
                            # insert register datas to next index
                            regs_org[0] = offs + r_desc[1] if split==0 else offs+32
                            regs_org[1] = size - r_desc[1] if split==0 else size-32
                            regs.desc[idx].insert(r_idx+1, regs_org)
                            split = 1
                        else:
                            split = 0

    def make_reg_list(self, axi_reg, lists, order=0, size_chk=0):
        ## Step1 - Match Debugging Register
        for data in self.datas:
            if axi_reg:
                if data[1].strip() in lists: ## Name
                    if not self.RegList.exist_name(data[1].strip(), data[5].strip()):
                        self.RegList.append(data)
            else:
                if (data[1].strip() in lists) or (GEN_ALL_REGSTER in lists): ## Name
                    self.RegList.append(data)
        ## Step2 - convert data to make script easily
        self.convert_datas(not (axi_reg or order)) # axi_reg : ascending order

        ## Step3[optional] - Split register list if size exceeds 32 bits
        if size_chk == 1:
            self.split_register_list()
        self.RegList.show()

    def generateTemplate_by_data(self, tmp_f, out_f, datas, lists=[], base_n="", base_addr=""):
        open(out_f, 'w').write(
            Template(filename=tmp_f).render(
            datas = datas,
            lists = lists,
            base_name = base_n,
            base_addr = base_addr,
            ).replace('\r\n', '\n').replace("[NEWLINE]\n", "")
        )

    def make_cmm(self, out_f, base_addr):
        global generated_f
        self.generateTemplate_by_data(
                                    tmp_f       = "template/fail_reg_cmm.mako",
                                    out_f       = "/".join([generated_f, out_f]),
                                    datas       = self.RegList.r_lists,
                                    base_addr   = base_addr,
                                    )

    def make_code(self, mako_f, out_f, base_name, base_addr, lists):
        global generated_f
        l_RegList = dict()
        for module, regs in self.RegList.r_lists.items():
            if (module in lists) or (GEN_ALL_REGSTER in lists):
                l_RegList[module] = regs
        self.generateTemplate_by_data(
                                    tmp_f       = "template/"+mako_f,
                                    out_f       = "/".join([generated_f, out_f]),
                                    datas       = l_RegList,
                                    lists       = lists,
                                    base_n      = base_name,
                                    base_addr   = base_addr,
                                    )

if __name__ == '__main__':
    arg = parser.parse_args()
    target_sys = arg.target.upper()
    target_dev = arg.dev
    generated_f = DEV_T[target_dev].lower()
    Path(generated_f).mkdir(parents=True, exist_ok=True)
    g_dev = DEV_T[target_dev]
    set_pcie_conf_value()

    ## Set configuration of device ##
    CL = ConfigList(target_dev)
    for key, value in CONFIG_LIST[DEV_T[target_dev]].items():
        CL.append_list(key, value)
    # CL.show()

    ## Get register map from excel
    SM = ScriptMaker()
    SM.create_split_datas(CL.reg_map_f)

    ## Parsing datas to generate header file and cmm script
    for key, config in CL.lists.items():
        print(f"Parsing for {key}")
        gen_list = GEN_ALL_REGSTER if config.gen_part == GEN_ALL_REGSTER else config.gen_lists
        SM.get_datas(config.gen_sheet_n)
        SM.set_reg_offset_recalculation(4) if key=="PHY_REG_CODE" else print("Skip recalculate for reg offset")
        if key=="CTRL_HIF_CODE":
            SM.make_reg_list(config.axi, gen_list, size_chk=1)
        elif key=="PHY_REG_CODE":
            SM.make_reg_list(config.axi, gen_list, 1)
        else:
            SM.make_reg_list(config.axi, gen_list)
        if "CMM" in key:
            if "CTRL_HIF" in key:
                SM.RegList.remove_white_space_of_desc_normal()
            else:
                SM.RegList.remove_white_space_of_desc()
            SM.make_cmm(config.gen_file_n, config.base_address)
        else:
            SM.make_code(config.mako_file_n,
                        config.gen_file_n,
                        config.name,
                        config.base_address,
                        gen_list)
        SM.reg_list_clear()

    print("generator tools are reached in endline!!")